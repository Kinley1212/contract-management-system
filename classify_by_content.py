"""
對 contract_list.xlsx 中未分類的文件，掃描文件內容搜索公司名稱關鍵詞，
只輸出 Excel，不做任何複製操作。

不覆蓋 process_contracts.py 的任何邏輯。
"""

from pathlib import Path
import pdfplumber
from docx import Document
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import MergedCell

DESKTOP = Path("/Users/fmpmac/Desktop")
CONTRACT_DIR = DESKTOP / "Contract"
SRC_EXCEL = DESKTOP / "contract_list.xlsx"
OUTPUT_DIR = DESKTOP / "contract_output"
OUT_EXCEL = OUTPUT_DIR / "contract_list_v2.xlsx"
REVIEW_EXCEL_V2 = OUTPUT_DIR / "contract_review_v2.xlsx"

# Read full document text (no limit)
SCAN_LIMIT = None

# Company name keywords to search in content
# Map: keyword → Contract company folder name
CONTENT_KEYWORDS = {
    # Delia
    "Delia Limited": "Delia Limited",
    "DELIA LIMITED": "Delia Limited",
    "Delia Ltd": "Delia Limited",
    # FM Investment
    "FM Investment Limited": "FM Investment Limited",
    "FM INVESTMENT LIMITED": "FM Investment Limited",
    "FM Investment Ltd": "FM Investment Limited",
    # Film Mall Entertainment
    "Film Mall Entertainment Limited": "Film Mall Entertainment Limited",
    "FILM MALL ENTERTAINMENT": "Film Mall Entertainment Limited",
    # Film Mall Limited
    "Film Mall Limited": "Film Mall Limited",
    "FILM MALL LIMITED": "Film Mall Limited",
    "Film Mall Ltd": "Film Mall Limited",
    # FM Group
    "FM Group (Holdings) Ltd": "FM Group (Holdings) Ltd",
    "FM GROUP (HOLDINGS)": "FM Group (Holdings) Ltd",
    "FM Group Holdings": "FM Group (Holdings) Ltd",
    # FM Event
    "FM Event Limited": "FM Event Limited",
    "FM EVENT LIMITED": "FM Event Limited",
    "FM Event Ltd": "FM Event Limited",
    # Film Mall Production
    "Film Mall Production Limited": "Film Mall Production Limited",
    "FILM MALL PRODUCTION": "Film Mall Production Limited",
    # Film Mall Producao
    "Film Mall Producao": "Film Mall Producao Limitada ",
    "FILM MALL PRODUCAO": "Film Mall Producao Limitada ",
    # FM Projects Limited
    "FM Projects Limited": "FM Projects Limited",
    "FM PROJECTS LIMITED": "FM Projects Limited",
    # FM Projects (HK)
    "FM Projects (HK) Limited": "FM Projects (HK) Limited",
    "FM PROJECTS (HK)": "FM Projects (HK) Limited",
    # FM Telemedia
    "FM Telemedia Limited": "FM Telemedia Limited",
    "FM TELEMEDIA LIMITED": "FM Telemedia Limited",
    # FUN FUN CHANGE
    "FUN FUN CHANGE": "FUN FUN CHANGE CO",
    "Fun Fun Change": "FUN FUN CHANGE CO",
    # Film Mall (SZ)
    "Film Mall (SZ)": "Film Mall (SZ) Ltd",
    "FILM MALL (SZ)": "Film Mall (SZ) Ltd",
    "深圳影市堂": "Film Mall (SZ) Ltd",
    # Love Smart
    "Love Smart": "Love Smart",
    "LOVE SMART": "Love Smart",
    # Tien River
    "TIEN RIVER": "TIEN RIVER - BVI",
    "Tien River": "TIEN RIVER - BVI",
    # 2P Entertainment
    "2P Entertainment": "2P Entertainment (Macau) Ltd",
    "2P ENTERTAINMENT": "2P Entertainment (Macau) Ltd",
    # 2P Workshop
    "2P Workshop": "2P Workshop",
    "2P WORKSHOP": "2P Workshop",
    # Vanuatu
    "Vanuatu": "Vanuatu",
    "VANUATU": "Vanuatu",
    # Chinese names (花蜜 = FM)
    "花蜜投資": "FM Investment Limited",
    "花蜜活動": "FM Event Limited",
    "花蜜事件": "FM Event Limited",
    "花蜜集團": "FM Group (Holdings) Ltd",
    "花蜜項目策劃（香港）": "FM Projects (HK) Limited",
    "花蜜項目策劃(香港)": "FM Projects (HK) Limited",
    "花蜜項目策劃": "FM Projects Limited",
    "花蜜項目": "FM Projects Limited",
    "花蜜電訊": "FM Telemedia Limited",
    "花蜜傳媒": "FM Telemedia Limited",
    "花蜜娛樂": "Film Mall Entertainment Limited",
    "花蜜製作": "Film Mall Production Limited",
}


def get_company_names(contract_dir: Path) -> set:
    return {p.name for p in contract_dir.iterdir() if p.is_dir()}


def extract_text_full(file_path: Path, limit=SCAN_LIMIT) -> str:
    try:
        if file_path.suffix.lower() == ".pdf":
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += (page.extract_text() or "")
            return text if limit is None else text[:limit]
        elif file_path.suffix.lower() == ".docx":
            doc = Document(file_path)
            text = "\n".join(p.text for p in doc.paragraphs)
            return text if limit is None else text[:limit]
    except Exception:
        return ""
    return ""


def search_companies_in_content(text: str, company_names: set) -> list:
    found = []
    for keyword, company_name in CONTENT_KEYWORDS.items():
        if company_name not in company_names:
            continue
        if keyword in text:
            if company_name not in found:
                found.append(company_name)
    return found


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def autowidth(ws):
    for col in ws.columns:
        col_letter = next((c.column_letter for c in col if not isinstance(c, MergedCell)), None)
        if not col_letter:
            continue
        max_len = max(
            (len(str(c.value)) for c in col if c.value and not isinstance(c, MergedCell)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main():
    company_names = get_company_names(CONTRACT_DIR)

    wb_src = openpyxl.load_workbook(SRC_EXCEL)
    ws_src = wb_src.active
    headers = [cell.value for cell in ws_src[1]]

    classified_rows = []
    unclassified_rows = []
    yellow_rows = []

    for row in ws_src.iter_rows(min_row=2):
        bg = row[0].fill.fgColor.rgb if row[0].fill and row[0].fill.fgColor else ""
        is_yellow = "FFFF00" in bg
        is_colored = bg not in ("", "00000000", "FFFFFFFF")
        vals = [cell.value for cell in row]
        if is_yellow:
            yellow_rows.append((vals, bg))
        elif is_colored:
            unclassified_rows.append((vals, bg))
        else:
            classified_rows.append(vals)

    print(f"已分類：{len(classified_rows)} 筆")
    print(f"多重匹配（黃）：{len(yellow_rows)} 筆")
    print(f"未分類：{len(unclassified_rows)} 筆，開始內容掃描...")

    newly_classified = []
    still_unclassified = []
    multi_content = []

    for i, (vals, orig_bg) in enumerate(unclassified_rows, 1):
        if i % 200 == 0:
            print(f"  掃描中... {i}/{len(unclassified_rows)}")

        original_path = vals[1]
        file_path = Path(str(original_path)) if original_path else None

        if not file_path or not file_path.exists():
            still_unclassified.append((vals, orig_bg))
            continue

        text = extract_text_full(file_path, SCAN_LIMIT)
        found = search_companies_in_content(text, company_names)

        if len(found) == 0:
            still_unclassified.append((vals, orig_bg))
        elif len(found) == 1:
            company = found[0]
            new_vals = [vals[0], vals[1], company, "", text[:300], "✅ 內容掃描建議分類"]
            newly_classified.append(new_vals)
        else:
            note = f"⚠️ 內容中出現多個公司：{', '.join(found)}"
            new_vals = vals[:2] + ["多重匹配", "", text[:300], note]
            multi_content.append((new_vals, "FFFF00"))

    print(f"\n內容掃描結果（僅輸出 Excel，未複製任何文件）：")
    print(f"  ✅ 內容掃描建議分類：{len(newly_classified)} 筆")
    print(f"  ⚠️  內容多重匹配：   {len(multi_content)} 筆")
    print(f"  ❌ 仍未分類：        {len(still_unclassified)} 筆")

    # --- Build full v2 Excel ---
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "合約清單v2"
    ws_out.append(headers)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    for vals in classified_rows:
        ws_out.append(vals)

    for vals in newly_classified:
        ws_out.append(vals)

    for vals, bg in yellow_rows:
        ws_out.append(vals)
        fill = make_fill(bg if bg not in ("", "00000000") else "FFFF00")
        for cell in ws_out[ws_out.max_row]:
            cell.fill = fill

    for vals, bg in multi_content:
        ws_out.append(vals)
        for cell in ws_out[ws_out.max_row]:
            cell.fill = make_fill("FFFF00")

    for vals, bg in still_unclassified:
        ws_out.append(vals)
        fill = make_fill(bg if bg not in ("", "00000000") else "FFCCCC")
        for cell in ws_out[ws_out.max_row]:
            cell.fill = fill

    autowidth(ws_out)
    wb_out.save(OUT_EXCEL)
    print(f"\n📄 全量 Excel 已保存：{OUT_EXCEL}")

    # --- Build review v2 ---
    wb_r = openpyxl.Workbook()
    ws_r = wb_r.active
    ws_r.title = "待確認清單v2"

    def section_header(title):
        ws_r.append([title])
        r = ws_r.max_row
        ws_r.cell(r, 1).font = Font(bold=True, size=12)
        ws_r.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))

    def write_section(rows_with_bg, fill_override=None):
        ws_r.append(headers)
        for cell in ws_r[ws_r.max_row]:
            cell.font = Font(bold=True)
        for vals, bg in rows_with_bg:
            ws_r.append(vals)
            color = fill_override or (bg if bg not in ("", "00000000") else "FFCCCC")
            f = make_fill(color)
            for cell in ws_r[ws_r.max_row]:
                cell.fill = f

    all_yellow = yellow_rows + multi_content
    section_header(f"▼ 多重匹配（請確認）— 共 {len(all_yellow)} 筆")
    write_section(all_yellow, fill_override="FFFF00")

    ws_r.append([])

    section_header(f"▼ 未分類（仍無法匹配）— 共 {len(still_unclassified)} 筆")
    write_section(still_unclassified)

    autowidth(ws_r)
    wb_r.save(REVIEW_EXCEL_V2)
    print(f"📋 待確認 Excel 已保存：{REVIEW_EXCEL_V2}")


if __name__ == "__main__":
    main()
