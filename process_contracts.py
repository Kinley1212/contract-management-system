import shutil
from pathlib import Path
import pdfplumber
from docx import Document
import openpyxl
from openpyxl.styles import PatternFill, Font

DESKTOP = Path("/Users/fmpmac/Desktop")
KUAHCHAI_DIR = DESKTOP / "kuahchai"
CONTRACT_DIR = DESKTOP / "Contract"
OUTPUT_EXCEL = DESKTOP / "contract_list.xlsx"
REVIEW_EXCEL = DESKTOP / "contract_review.xlsx"
CONTENT_LIMIT = 300

# Paths to force-unclassify regardless of folder names
EXCLUDED_PATHS = [
    KUAHCHAI_DIR / "CURRENT" / "EVENT" / "2026黃大仙情人節新春市集" / "2P",
]

# Alias map: folder name → Contract company name
ALIAS_MAP = {
    # Delia
    "DELIA": "Delia Limited",
    "Delia Ltd": "Delia Limited",
    # FM Investment
    "FMI": "FM Investment Limited",
    "FM INVESTMENT": "FM Investment Limited",
    "FM INVESTMENT LTD--1": "FM Investment Limited",
    "FM Investment Ltd": "FM Investment Limited",
    "FM Investment": "FM Investment Limited",
    "FM Investment-Breakdown": "FM Investment Limited",
    "FMI - Breakdown": "FM Investment Limited",
    # Film Mall Entertainment
    "FME": "Film Mall Entertainment Limited",
    "fme": "Film Mall Entertainment Limited",
    "FM Entertainment": "Film Mall Entertainment Limited",
    "FME(花蜜深圳)": "Film Mall Entertainment Limited",
    # Film Mall Limited
    "FILM MALL LTD": "Film Mall Limited",
    "Film Mall Ltd": "Film Mall Limited",
    "Film Mall Limitada": "Film Mall Limited",
    "FML Breakdown-1": "Film Mall Limited",
    "FILM MALL": "Film Mall Limited",
    # FM Group
    "FM GROUP": "FM Group (Holdings) Ltd",
    "fmg": "FM Group (Holdings) Ltd",
    "FM GROUP SZ": "FM Group (Holdings) Ltd",
    # FM Event
    "FM EVENT LTD": "FM Event Limited",
    "FM EVENT LTD - 複製": "FM Event Limited",
    "FM Event": "FM Event Limited",
    "EVENT": "FM Event Limited",
    # Film Mall Production
    "FM PRODUCTION": "Film Mall Production Limited",
    "FM Production": "Film Mall Production Limited",
    "FM Production Ltd": "Film Mall Production Limited",
    "Film Mall Production (Macau)": "Film Mall Producao Limitada ",
    # FM Projects Limited
    "FM PROJECTS": "FM Projects Limited",
    "FM Projects": "FM Projects Limited",
    "FMP": "FM Projects Limited",
    "FMPL": "FM Projects Limited",
    "FMPL Breakdown": "FM Projects Limited",
    "Expenses - FM PROJECTS": "FM Projects Limited",
    # FM Projects (HK) Limited
    "FMP HK": "FM Projects (HK) Limited",
    "FMP(HK)": "FM Projects (HK) Limited",
    "FMP-HK": "FM Projects (HK) Limited",
    "FMP-MD-Breakdown": "FM Projects (HK) Limited",
    # FM Telemedia
    "FM TELEMEDIA": "FM Telemedia Limited",
    "fmt": "FM Telemedia Limited",
    "FMT  Breakdown": "FM Telemedia Limited",
    # FUN FUN CHANGE
    "FUN-FUN CHANGE": "FUN FUN CHANGE CO",
    # Love Smart
    "LOVE SMART": "Love Smart",
    # Tien River — WNT 是 FM Investment 下的項目文件夾，不映射公司
    "Tien River": "TIEN RIVER - BVI",
    "WNT-Tien River": "TIEN RIVER - BVI",
    "ENTERTAINMENT-BVI": "TIEN RIVER - BVI",
    # Film Mall (SZ)
    "fm (sz)": "Film Mall (SZ) Ltd",
    "FM Group SZ": "Film Mall (SZ) Ltd",
}

# Rotating color palette for unclassified groups (by parent folder name)
UNCLASSIFIED_PALETTE = [
    "FFD9B3", "D9FFD9", "D9D9FF", "FFD9F5", "D9FFFF",
    "FFFFD9", "F5D9FF", "D9F5FF", "FFE8D9", "D9FFE8",
    "E8FFD9", "FFD9D9", "D9E8FF", "FFD9E8", "E8D9FF",
]


def extract_text(file_path: Path) -> str:
    try:
        if file_path.suffix.lower() == ".pdf":
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += (page.extract_text() or "")
                    if len(text) >= CONTENT_LIMIT:
                        break
            return text[:CONTENT_LIMIT]
        elif file_path.suffix.lower() == ".docx":
            doc = Document(file_path)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
                if len(text) >= CONTENT_LIMIT:
                    break
            return text[:CONTENT_LIMIT]
    except Exception as e:
        return f"[提取失敗: {e}]"
    return ""


def get_company_folders(contract_dir: Path) -> dict:
    return {p.name: p for p in contract_dir.iterdir() if p.is_dir()}


def is_excluded(file_path: Path) -> bool:
    for excl in EXCLUDED_PATHS:
        try:
            file_path.relative_to(excl)
            return True
        except ValueError:
            pass
    return False


def resolve_company(file_path: Path, companies: dict) -> tuple:
    """
    Returns (matched_company_or_None, is_multi_match).
    Uses 'last company in path' rule when multiple matches found.
    """
    rel = file_path.relative_to(KUAHCHAI_DIR)
    ancestors = list(rel.parts[:-1])  # folder names only, exclude filename

    matches_with_depth = []  # (depth_index, company_name)

    for i, ancestor in enumerate(ancestors):
        company = None
        if ancestor in ALIAS_MAP:
            candidate = ALIAS_MAP[ancestor]
            if candidate in companies:
                company = candidate
        if company is None:
            for cname in companies:
                if cname in ancestor or ancestor in cname:
                    company = cname
                    break
        if company:
            matches_with_depth.append((i, company))

    if not matches_with_depth:
        return None, False

    # Deduplicate: keep the deepest occurrence of each company
    seen = {}
    for depth, company in matches_with_depth:
        seen[company] = depth

    unique_companies = list(seen.keys())
    is_multi = len(unique_companies) > 1

    # Pick the company with the greatest depth (last in path)
    best_company = max(seen.items(), key=lambda x: x[1])[0]
    return best_company, is_multi


def safe_dest(dest_dir: Path, filename: str) -> Path:
    dest_path = dest_dir / filename
    if dest_path.exists():
        stem = Path(filename).stem
        suffix = Path(filename).suffix
        counter = 1
        while dest_path.exists():
            dest_path = dest_dir / f"{stem}_{counter}{suffix}"
            counter += 1
    return dest_path


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def apply_row_fill(ws, row_num: int, fill: PatternFill):
    for cell in ws[row_num]:
        cell.fill = fill


def autowidth(ws):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        col_letter = next((c.column_letter for c in col if not isinstance(c, MergedCell)), None)
        if col_letter is None:
            continue
        max_len = max(
            (len(str(c.value)) for c in col if c.value and not isinstance(c, MergedCell)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main():
    if not KUAHCHAI_DIR.exists():
        print(f"[錯誤] 找不到源文件夾: {KUAHCHAI_DIR}")
        return
    if not CONTRACT_DIR.exists():
        print(f"[錯誤] 找不到目標文件夾: {CONTRACT_DIR}")
        return

    # Auto-clean Contract subfolders before each run
    deleted = sum(1 for f in CONTRACT_DIR.rglob("*") if f.is_file() and not f.unlink())
    for name in ["contract_list.xlsx", "contract_review.xlsx"]:
        p = DESKTOP / name
        if p.exists():
            p.unlink()
    print(f"已清空 Contract 內 {deleted} 個舊文件及舊 Excel")

    companies = get_company_folders(CONTRACT_DIR)
    print(f"找到 {len(companies)} 個公司文件夾")

    target_files = [
        p for p in KUAHCHAI_DIR.rglob("*")
        if p.is_file() and p.suffix.lower() in (".pdf", ".docx")
    ]
    print(f"找到 {len(target_files)} 個文件（.pdf / .docx）")

    # --- Build full Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合約清單"
    headers = ["文件名", "kuahchai 原始路徑", "匹配公司", "複製後路徑", "前300字內容", "備註"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    yellow_fill = make_fill("FFFF00")
    red_fill = make_fill("FFCCCC")

    # Color map for unclassified parent folders
    unclassified_color_map: dict[str, PatternFill] = {}
    palette_index = 0

    # Dedup tracker: company_name → set of filenames already copied
    copied_tracker: dict[str, set] = {c: set() for c in companies}

    count_copied = 0
    count_unclassified = 0
    count_multi = 0
    count_dup = 0

    for i, file_path in enumerate(target_files, 1):
        if i % 500 == 0:
            print(f"  處理中... {i}/{len(target_files)}")

        filename = file_path.name
        original_path = str(file_path)
        parent_folder = file_path.parent.name

        # --- Excluded path → red unclassified ---
        if is_excluded(file_path):
            content = extract_text(file_path)
            if parent_folder not in unclassified_color_map:
                unclassified_color_map[parent_folder] = make_fill(
                    UNCLASSIFIED_PALETTE[palette_index % len(UNCLASSIFIED_PALETTE)]
                )
                palette_index += 1
            ws.append([filename, original_path, "未分類（排除路徑）", "", content, "2026黃大仙情人節新春市集/2P 排除"])
            apply_row_fill(ws, ws.max_row, unclassified_color_map[parent_folder])
            count_unclassified += 1
            continue

        company, is_multi = resolve_company(file_path, companies)
        content = extract_text(file_path)

        if company is None:
            # Unclassified: color by parent folder
            if parent_folder not in unclassified_color_map:
                unclassified_color_map[parent_folder] = make_fill(
                    UNCLASSIFIED_PALETTE[palette_index % len(UNCLASSIFIED_PALETTE)]
                )
                palette_index += 1
            ws.append([filename, original_path, "未分類", "", content, ""])
            apply_row_fill(ws, ws.max_row, unclassified_color_map[parent_folder])
            count_unclassified += 1

        else:
            # Dedup check
            if filename in copied_tracker[company]:
                ws.append([filename, original_path, company, "", content, "⏭ 重複文件，已跳過"])
                count_dup += 1
                continue

            dest_path = safe_dest(companies[company], filename)
            shutil.copy2(file_path, dest_path)
            copied_tracker[company].add(filename)

            note = "⚠️ 多重匹配，已自動取最深層公司" if is_multi else ""
            fill = yellow_fill if is_multi else None
            ws.append([filename, original_path, company, str(dest_path), content, note])
            if fill:
                apply_row_fill(ws, ws.max_row, fill)
            count_copied += 1
            if is_multi:
                count_multi += 1

    autowidth(ws)
    wb.save(OUTPUT_EXCEL)
    print(f"✅ 全量 Excel 已保存：{OUTPUT_EXCEL}")

    # --- Build review Excel (yellow first, then unclassified) ---
    wb2 = openpyxl.load_workbook(OUTPUT_EXCEL)
    ws2 = wb2.active
    wb_r = openpyxl.Workbook()
    ws_r = wb_r.active
    ws_r.title = "待確認清單"

    def section_header(ws, title):
        ws.append([title])
        row = ws.max_row
        ws.cell(row, 1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    def copy_rows(ws, rows_data):
        for vals, src_fill in rows_data:
            ws.append(vals)
            new_row = ws.max_row
            for cell in ws[new_row]:
                cell.fill = PatternFill(
                    start_color=src_fill.fgColor.rgb,
                    end_color=src_fill.fgColor.rgb,
                    fill_type="solid"
                )

    yellow_rows = []
    unclassified_rows = []

    for row in ws2.iter_rows(min_row=2):
        bg = row[0].fill.fgColor.rgb if row[0].fill and row[0].fill.fgColor else ""
        is_yellow = "FFFF00" in bg
        is_colored = bg not in ("", "00000000", "FFFFFFFF")
        if not (is_yellow or is_colored):
            continue
        vals = [cell.value for cell in row]
        src_fill = row[0].fill
        if is_yellow:
            yellow_rows.append((vals, src_fill))
        else:
            unclassified_rows.append((vals, src_fill))

    # Section 1: Multi-match
    section_header(ws_r, f"▼ 多重匹配（自動解析，請確認）— 共 {len(yellow_rows)} 筆")
    ws_r.append(headers)
    for cell in ws_r[ws_r.max_row]:
        cell.font = Font(bold=True)
    copy_rows(ws_r, yellow_rows)

    ws_r.append([])  # blank separator

    # Section 2: Unclassified
    section_header(ws_r, f"▼ 未分類（無法自動匹配）— 共 {len(unclassified_rows)} 筆")
    ws_r.append(headers)
    for cell in ws_r[ws_r.max_row]:
        cell.font = Font(bold=True)
    copy_rows(ws_r, unclassified_rows)

    autowidth(ws_r)
    wb_r.save(REVIEW_EXCEL)
    print(f"📋 待確認 Excel 已保存：{REVIEW_EXCEL}")

    print(f"\n{'='*40}")
    print(f"✅ 成功複製：     {count_copied} 份")
    print(f"⚠️  自動解析多匹配：{count_multi} 份")
    print(f"⏭  跳過重複文件：  {count_dup} 份")
    print(f"❌ 未分類：       {count_unclassified} 份")
    print(f"{'='*40}")


if __name__ == "__main__":
    main()
